"""
机器人场地竞赛 - 最优策略设计
==============================================
题目：10×10网格，机器人1m/s，食物存活3秒
目标：3小时内最大化积分

Interface: importing this module only loads the dataset, the precomputed
maps and the strategy functions (strategy_* / make_weighted_strategy);
the full analysis runs via run_full_analysis() when executed as a script.
testenv/learn_blackbox.py wraps these strategies for the black-box test
environment (testenv/env.py).
Coordinates: this module uses 0-indexed (r, c) where r is the FIRST number
in the dataset position "(3,1)"; the environment protocol uses 1-indexed
(x, y), so (r, c) = (x-1, y-1).
"""

import os
import openpyxl
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import rcParams
rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
rcParams['axes.unicode_minus'] = False
from collections import Counter, defaultdict
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# 1. 数据加载
# ============================================================
print("=" * 60)
print("1. 数据加载")
print("=" * 60)

# data.xlsx is a byte-identical copy of the GBK-named attachment (the raw
# name cannot be opened via a UTF-8 path on this system)
_DATA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.xlsx')
wb = openpyxl.load_workbook(_DATA_PATH, data_only=True)
ws = wb['Sheet1']
data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    data.append(list(row))

times = np.array([d[0] for d in data])
values = np.array([d[2] for d in data])

rows_list, cols_list = [], []
for d in data:
    p = str(d[1]).strip('()')
    r, c = p.split(',')
    rows_list.append(int(r))
    cols_list.append(int(c))
rows_arr = np.array(rows_list)
cols_arr = np.array(cols_list)

N = len(data)
TOTAL_TIME = 10800  # 3 hours
GRID_SIZE = 10
FOOD_LIFETIME = 3  # seconds
ROBOT_SPEED = 1  # m/s

print(f"总食物数: {N}")
print(f"时间跨度: {min(times):.0f} - {max(times):.0f} 秒")
print(f"分值: min={min(values)}, max={max(values)}, mean={values.mean():.2f}, median={np.median(values):.0f}")

# ============================================================
# 2. 时间规律分析
# ============================================================
print("\n" + "=" * 60)
print("2. 时间规律分析")
print("=" * 60)

times_sorted = np.sort(times)
inter_arrival = np.diff(times_sorted)
print(f"到达间隔: min={inter_arrival.min():.0f}s, max={inter_arrival.max():.0f}s, "
      f"mean={inter_arrival.mean():.2f}s, std={inter_arrival.std():.2f}s")

# 分时段统计
hour_labels = ['第1小时', '第2小时', '第3小时']
for h in range(3):
    mask = (times >= h*3600) & (times < (h+1)*3600)
    print(f"  {hour_labels[h]}: {mask.sum()}个食物, 均值={values[mask].mean():.2f}, "
          f"总积分={values[mask].sum():.0f}")

# 每10分钟统计
print("\n每10分钟食物数量与总积分:")
for i in range(18):
    start, end = i*600, (i+1)*600
    mask = (times >= start) & (times < end)
    print(f"  {start//60:3d}-{end//60:3d}min: {mask.sum():2d}个, 总积分{values[mask].sum():.0f}")

# ============================================================
# 3. 空间规律分析
# ============================================================
print("\n" + "=" * 60)
print("3. 空间规律分析")
print("=" * 60)

# 3a. 每个位置的出现频率
grid_count = np.zeros((GRID_SIZE, GRID_SIZE))
grid_value_sum = np.zeros((GRID_SIZE, GRID_SIZE))
for i in range(N):
    r, c = rows_arr[i] - 1, cols_arr[i] - 1
    grid_count[r, c] += 1
    grid_value_sum[r, c] += values[i]

grid_avg_value = np.divide(grid_value_sum, grid_count, where=grid_count > 0)

print("\n各位置食物出现次数热力图 (10×10):")
for r in range(GRID_SIZE):
    row_str = ' '.join(f'{int(grid_count[r,c]):4d}' for c in range(GRID_SIZE))
    print(f"  行{r+1:2d}: {row_str}")

print("\n各位置平均分值热力图 (10×10):")
for r in range(GRID_SIZE):
    row_str = ' '.join(f'{grid_avg_value[r,c]:5.1f}' for c in range(GRID_SIZE))
    print(f"  行{r+1:2d}: {row_str}")

# 3b. 行列边际分布
print("\n列边际分布:")
for c in range(GRID_SIZE):
    col_count = grid_count[:, c].sum()
    col_avg = grid_value_sum[:, c].sum() / col_count if col_count > 0 else 0
    print(f"  列{c+1}: {int(col_count):4d}次, 均值{col_avg:.2f}")

print("\n行边际分布:")
for r in range(GRID_SIZE):
    row_count = grid_count[r, :].sum()
    row_avg = grid_value_sum[r, :].sum() / row_count if row_count > 0 else 0
    print(f"  行{r+1}: {int(row_count):4d}次, 均值{row_avg:.2f}")

# ============================================================
# 4. 分值分布分析
# ============================================================
print("\n" + "=" * 60)
print("4. 分值分布分析")
print("=" * 60)

val_bins = [0, 3, 5, 10, 15, 20, 25, 30, 41]
print(f"\n分值区间分布:")
for i in range(len(val_bins)-1):
    mask = (values >= val_bins[i]) & (values < val_bins[i+1])
    print(f"  [{val_bins[i]:2d}, {val_bins[i+1]:2d}): {mask.sum():4d}个 ({mask.sum()/N*100:5.1f}%)")

# 高分值食物空间分布
high_val_mask = values >= 20
print(f"\n高分值食物(≥20)空间分布 (共{high_val_mask.sum()}个):")
high_grid = np.zeros((GRID_SIZE, GRID_SIZE))
for i in np.where(high_val_mask)[0]:
    high_grid[rows_arr[i]-1, cols_arr[i]-1] += 1
for r in range(GRID_SIZE):
    print(' '.join(f'{int(high_grid[r,c]):3d}' for c in range(GRID_SIZE)))

# ============================================================
# 5. 可达性分析 (核心)
# ============================================================
print("\n" + "=" * 60)
print("5. 可达性分析 (曼哈顿距离≤3)")
print("=" * 60)

def manhattan(r1, c1, r2, c2):
    return abs(r1 - r2) + abs(c1 - c2)

# 5a. 每个位置的"覆盖范围"内有多少食物
print("\n各位置的3步可达食物总数:")
coverage_count = np.zeros((GRID_SIZE, GRID_SIZE))
coverage_value = np.zeros((GRID_SIZE, GRID_SIZE))
for r in range(GRID_SIZE):
    for c in range(GRID_SIZE):
        total_v = 0
        total_n = 0
        for i in range(N):
            d = manhattan(r, c, rows_arr[i]-1, cols_arr[i]-1)
            if d <= 3:
                total_n += 1
                total_v += values[i]
        coverage_count[r, c] = total_n
        coverage_value[r, c] = total_v

for r in range(GRID_SIZE):
    row_str = ' '.join(f'{int(coverage_count[r,c]):5d}' for c in range(GRID_SIZE))
    print(f"  行{r+1:2d}: {row_str}")

print("\n各位置的3步可达食物总积分:")
for r in range(GRID_SIZE):
    row_str = ' '.join(f'{coverage_value[r,c]:6.0f}' for c in range(GRID_SIZE))
    print(f"  行{r+1:2d}: {row_str}")

# 5b. 最佳等待位置
best_pos = np.unravel_index(coverage_value.argmax(), coverage_value.shape)
print(f"\n最佳等待位置(覆盖总积分最大): ({best_pos[0]+1}, {best_pos[1]+1})")
print(f"  覆盖食物数: {coverage_count[best_pos]:.0f}")
print(f"  覆盖总积分: {coverage_value[best_pos]:.0f}")

# Top 10 等待位置
flat_indices = np.argsort(coverage_value.flatten())[::-1][:10]
print("\nTop 10 等待位置:")
for idx in flat_indices:
    r, c = idx // GRID_SIZE, idx % GRID_SIZE
    print(f"  ({r+1}, {c+1}): 覆盖{int(coverage_count[r,c])}个, 总积分{coverage_value[r,c]:.0f}")

# ============================================================
# 9. 策略仿真引擎
# ============================================================
print("\n" + "=" * 60)
print("9. 构建仿真引擎")
print("=" * 60)

class RobotSimulator:
    """机器人竞赛仿真器"""

    def __init__(self, times, rows, cols, values, food_lifetime=3, total_time=10800):
        self.times = times
        self.rows = rows
        self.cols = cols
        self.values = values
        self.food_lifetime = food_lifetime
        self.total_time = total_time
        self.N = len(times)

        # 按时间排序
        self.foods = sorted(zip(times, rows, cols, values), key=lambda x: x[0])

    def manhattan(self, r1, c1, r2, c2):
        return abs(r1 - r2) + abs(c1 - c2)

    def simulate(self, strategy, start_pos=(0, 0)):
        """
        执行仿真
        strategy: 决策函数 (robot_pos, active_foods, current_time) -> (target_r, target_c) or None
        """
        robot_r, robot_c = start_pos
        target_r, target_c = None, None  # 当前移动目标
        moving_to_food_idx = None  # 正在前往的食物索引 (-1表示无)
        total_score = 0
        eaten_count = 0
        missed_count = 0

        # 食物状态: -1=未出现, 0=已过期/被吃, 1=活跃
        food_status = np.zeros(self.N, dtype=int) - 1  # -1: not yet appeared

        # 事件驱动仿真
        food_idx = 0  # 下一个要出现的食物

        for t in range(self.total_time + 1):
            # 1. 新食物出现
            while food_idx < self.N and self.foods[food_idx][0] == t:
                food_status[food_idx] = 1  # active
                food_idx += 1

            # 2. 食物过期检查
            for i in range(food_idx):
                if food_status[i] == 1:
                    appear_t = self.foods[i][0]
                    if t >= appear_t + self.food_lifetime:
                        food_status[i] = 0  # expired
                        missed_count += 1

            # 3. 检查是否到达目标
            if moving_to_food_idx is not None and moving_to_food_idx >= 0:
                f_t, f_r, f_c, f_v = self.foods[moving_to_food_idx]
                if robot_r == f_r and robot_c == f_c:
                    # 到达食物位置，立即吃
                    if food_status[moving_to_food_idx] == 1:
                        total_score += f_v
                        eaten_count += 1
                        food_status[moving_to_food_idx] = -2  # eaten
                    moving_to_food_idx = None
                    target_r, target_c = None, None

            # 4. 决策：选择下一个目标
            if moving_to_food_idx is None:
                # 收集活跃食物
                active = []
                for i in range(food_idx):
                    if food_status[i] == 1:
                        f_t, f_r, f_c, f_v = self.foods[i]
                        remaining = self.food_lifetime - (t - f_t)
                        d = self.manhattan(robot_r, robot_c, f_r, f_c)
                        if d <= remaining:  # 可达
                            active.append((i, d, f_r, f_c, f_v, remaining))

                if active:
                    # 使用策略决策
                    choice = strategy(robot_r, robot_c, active, t)
                    if choice is not None:
                        idx, tr, tc = choice
                        moving_to_food_idx = idx
                        target_r, target_c = tr, tc

            # 5. 移动
            if target_r is not None:
                # 朝目标移动一步
                if robot_r < target_r:
                    robot_r += 1
                elif robot_r > target_r:
                    robot_r -= 1
                elif robot_c < target_c:
                    robot_c += 1
                elif robot_c > target_c:
                    robot_c -= 1

            # 6. 如果没有目标且没有在追食物，根据策略决定等待位置
            if target_r is None and moving_to_food_idx is None:
                wait_pos = strategy(robot_r, robot_c, [], t)  # idle decision
                if wait_pos is not None:
                    wr, wc = wait_pos
                    if wr != robot_r or wc != robot_c:
                        target_r, target_c = wr, wc

        return {
            'total_score': total_score,
            'eaten_count': eaten_count,
            'missed_count': missed_count,
            'score_per_minute': total_score / (self.total_time / 60),
            'score_per_hour': total_score / (self.total_time / 3600),
        }

    def simulate_discrete(self, strategy_func, start_pos=(0, 0)):
        """
        离散时间仿真 (每秒一步)，更精确
        strategy_func: function(robot_r, robot_c, active_foods_list, t) -> (target_r, target_c)
        active_foods_list: [(idx, food_r, food_c, value, remaining_time), ...]
        返回 (target_r, target_c) 表示移动目标，返回 None 表示原地等待
        """
        robot_r, robot_c = start_pos
        target_r, target_c = start_pos
        total_score = 0
        eaten_foods = set()
        path = [(0, robot_r, robot_c)]  # (t, r, c)

        for t in range(self.total_time + 1):
            # 当前活跃食物
            active = []
            for i, (f_t, f_r, f_c, f_v) in enumerate(self.foods):
                if i in eaten_foods:
                    continue
                if f_t <= t < f_t + self.food_lifetime:
                    remaining = self.food_lifetime - (t - f_t)
                    active.append((i, f_r, f_c, f_v, remaining))

            # 检查是否在某个食物上
            for i, f_r, f_c, f_v, rem in active:
                if robot_r == f_r and robot_c == f_c and i not in eaten_foods:
                    total_score += f_v
                    eaten_foods.add(i)
                    break

            # 决策
            target = strategy_func(robot_r, robot_c, active, t)
            if target is not None:
                target_r, target_c = target
            else:
                target_r, target_c = robot_r, robot_c  # 原地等待

            # 移动一步
            if robot_r < target_r:
                robot_r += 1
            elif robot_r > target_r:
                robot_r -= 1
            elif robot_c < target_c:
                robot_c += 1
            elif robot_c > target_c:
                robot_c -= 1

            if t % 3600 == 0 and t > 0:
                path.append((t, robot_r, robot_c))

        path.append((self.total_time, robot_r, robot_c))
        return {
            'total_score': int(total_score),
            'eaten_count': len(eaten_foods),
            'score_per_minute': total_score / (self.total_time / 60),
            'score_per_hour': total_score / (self.total_time / 3600),
            'path': path,
        }

# ============================================================
# 10. 策略定义
# ============================================================
print("\n" + "=" * 60)
print("10. 定义多种策略")
print("=" * 60)

# 预计算: 各位置的期望收益 (用于等待策略)
# 位置(r,c)的期望每秒收益 = sum(覆盖范围内的食物积分) / total_time
expected_value_map = coverage_value / TOTAL_TIME
best_wait_pos = np.unravel_index(coverage_value.argmax(), coverage_value.shape)

def strategy_static_best(robot_r, robot_c, active, t):
    """策略A: 始终待在最佳位置 (3,7)，有可达食物就去"""
    wr, wc = best_wait_pos
    if not active:
        return (wr, wc)
    # 选择可达且价值最高的
    best = None
    best_score = -1
    for idx, f_r, f_c, f_v, rem in active:
        d = abs(robot_r - f_r) + abs(robot_c - f_c)
        if d <= rem:
            score = f_v  # 可以加权 f_v / (d+1)
            if score > best_score:
                best_score = score
                best = (f_r, f_c)
    if best is None:
        return (wr, wc)
    return best

def strategy_greedy_value(robot_r, robot_c, active, t):
    """策略B: 贪心 - 选可达的最高价值食物"""
    if not active:
        return best_wait_pos
    best = None
    best_score = -1
    for idx, f_r, f_c, f_v, rem in active:
        d = abs(robot_r - f_r) + abs(robot_c - f_c)
        if d <= rem:
            score = f_v
            if score > best_score:
                best_score = score
                best = (f_r, f_c)
    return best if best else best_wait_pos

def strategy_value_per_distance(robot_r, robot_c, active, t):
    """策略C: 价值/距离比最大"""
    if not active:
        return best_wait_pos
    best = None
    best_ratio = -1
    for idx, f_r, f_c, f_v, rem in active:
        d = abs(robot_r - f_r) + abs(robot_c - f_c)
        if d <= rem:
            ratio = f_v / (d + 1)
            if ratio > best_ratio:
                best_ratio = ratio
                best = (f_r, f_c)
    return best if best else best_wait_pos

def strategy_adaptive_wait(robot_r, robot_c, active, t):
    """策略D: 自适应 - 有食物时比值决策，无食物时动态选择等待位置"""
    if not active:
        # 基于历史频率选择当前位置附近的最佳等待点
        # 考虑不需要移动太远的位置
        best_pos = best_wait_pos
        best_val = -1
        for r in range(GRID_SIZE):
            for c in range(GRID_SIZE):
                d = abs(robot_r - r) + abs(robot_c - c)
                # 折现期望收益
                discounted = expected_value_map[r, c] * np.exp(-0.05 * d)
                if discounted > best_val:
                    best_val = discounted
                    best_pos = (r, c)
        return best_pos

    # 有食物时用价值/距离比
    best = None
    best_ratio = -1
    for idx, f_r, f_c, f_v, rem in active:
        d = abs(robot_r - f_r) + abs(robot_c - f_c)
        if d <= rem and rem > 0:
            ratio = f_v / (d + 0.5)  # 略微偏向近距离
            if ratio > best_ratio:
                best_ratio = ratio
                best = (f_r, f_c)
    return best if best else (robot_r, robot_c)  # 无可达食物则不动

# 策略E: 基于全局优化的预计算策略
# 使用动态规划 / value iteration
print("预计算最优值函数...")

def compute_optimal_value_map(grid_count, grid_value_sum, food_lifetime=3, discount=1.0):
    """
    计算每个位置的"价值": 站在这里，能期望吃到多少积分的食物
    使用覆盖范围内的加权积分
    """
    value_map = np.zeros((GRID_SIZE, GRID_SIZE))
    for r in range(GRID_SIZE):
        for c in range(GRID_SIZE):
            total = 0
            for dr in range(-food_lifetime, food_lifetime+1):
                for dc in range(-food_lifetime, food_lifetime+1):
                    if abs(dr) + abs(dc) <= food_lifetime:
                        nr, nc = r + dr, c + dc
                        if 0 <= nr < GRID_SIZE and 0 <= nc < GRID_SIZE:
                            # 该位置的食物积分 * 到达概率
                            # 实际上如果机器人空闲就可以到达
                            total += grid_value_sum[nr, nc]
            value_map[r, c] = total
    return value_map

optimal_value_map = compute_optimal_value_map(grid_count, grid_value_sum)

def strategy_ml_optimal(robot_r, robot_c, active, t):
    """策略E: ML最优 - 基于全局价值图决策"""
    if not active:
        # 移动到最优位置
        best_pos = np.unravel_index(optimal_value_map.argmax(), optimal_value_map.shape)
        return best_pos

    # 有多个食物时，综合考虑
    candidates = []
    for idx, f_r, f_c, f_v, rem in active:
        d = abs(robot_r - f_r) + abs(robot_c - f_c)
        if d <= rem:
            # 吃完后该位置的价值(可以继续等)
            future_value = optimal_value_map[f_r, f_c] * 0.001  # 微小权重
            total_value = f_v + future_value
            candidates.append((total_value, d, f_r, f_c))

    if not candidates:
        return best_wait_pos

    # 选总价值最大的
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][2:]

# 优化策略C中的权重参数
def make_weighted_strategy(alpha, beta):
    """alpha: 价值权重指数, beta: 距离惩罚系数"""
    def strategy(robot_r, robot_c, active, t):
        if not active:
            return best_wait_pos
        best = None
        best_score = -1
        for idx, f_r, f_c, f_v, rem in active:
            d = abs(robot_r - f_r) + abs(robot_c - f_c)
            if d <= rem:
                score = (f_v ** alpha) / (d + beta)
                if score > best_score:
                    best_score = score
                    best = (f_r, f_c)
        return best if best else best_wait_pos
    return strategy

# ============================================================
# 17. 高级策略：时间感知 + 巡逻路线 + Q-learning
# ============================================================
print("\n" + "=" * 60)
print("17. 高级策略设计")
print("=" * 60)

# 17a. 时间感知策略：不同时间段食物热点可能不同
# 计算每个时段的最优等待位置
time_blocks = 6  # 每30分钟一个块
block_size = TOTAL_TIME // time_blocks
block_best_positions = []
for blk in range(time_blocks):
    start_t = blk * block_size
    end_t = (blk + 1) * block_size
    mask = (times >= start_t) & (times < end_t)
    blk_grid_val = np.zeros((GRID_SIZE, GRID_SIZE))
    for i in np.where(mask)[0]:
        r, c = rows_arr[i]-1, cols_arr[i]-1
        blk_grid_val[r, c] += values[i]
    blk_coverage = np.zeros((GRID_SIZE, GRID_SIZE))
    for r in range(GRID_SIZE):
        for c in range(GRID_SIZE):
            total = 0
            for dr in range(-3, 4):
                for dc in range(-3, 4):
                    if abs(dr) + abs(dc) <= 3:
                        nr, nc = r + dr, c + dc
                        if 0 <= nr < GRID_SIZE and 0 <= nc < GRID_SIZE:
                            total += blk_grid_val[nr, nc]
            blk_coverage[r, c] = total
    best = np.unravel_index(blk_coverage.argmax(), blk_coverage.shape)
    block_best_positions.append(best)
    print(f"  时段{blk+1} ({start_t//60}-{end_t//60}min): 最佳位置({best[0]+1},{best[1]+1}), 覆盖积分{blk_coverage[best]:.0f}")

def strategy_time_aware(robot_r, robot_c, active, t):
    """策略G: 时间感知 - 根据时段选择等待位置"""
    blk = min(t // block_size, time_blocks - 1)
    wr, wc = block_best_positions[blk]

    if not active:
        return (wr, wc)

    best = None
    best_score = -1
    for idx, f_r, f_c, f_v, rem in active:
        d = abs(robot_r - f_r) + abs(robot_c - f_c)
        if d <= rem:
            # 吃完后的位置价值
            blk2 = min((t + d) // block_size, time_blocks - 1)
            future_r, future_c = block_best_positions[blk2]
            future_dist = abs(f_r - future_r) + abs(f_c - future_c)
            score = f_v - 0.1 * future_dist  # 惩罚远离下一个等待位置
            if score > best_score:
                best_score = score
                best = (f_r, f_c)
    return best if best else (wr, wc)


# 17b. 巡逻策略：在热区循环移动
print("\n设计巡逻路线...")
# 基于覆盖价值图设计巡逻路线：在top热点之间循环
hot_positions = []
flat_idx = np.argsort(coverage_value.flatten())[::-1][:5]  # top 5
for idx in flat_idx:
    r, c = idx // GRID_SIZE, idx % GRID_SIZE
    hot_positions.append((r, c))

# 贪心TSP路径
patrol_route = [hot_positions[0]]
remaining = set(hot_positions[1:])
current = hot_positions[0]
while remaining:
    next_pos = min(remaining, key=lambda p: abs(current[0]-p[0]) + abs(current[1]-p[1]))
    patrol_route.append(next_pos)
    remaining.remove(next_pos)
    current = next_pos

print(f"巡逻路线: {' -> '.join(f'({r+1},{c+1})' for r,c in patrol_route)}")

def strategy_patrol(robot_r, robot_c, active, t):
    """策略H: 巡逻 - 在热点间循环移动，有食物时择优"""
    if not active:
        # 找到巡逻路线上的下一个目标
        cycle_pos = (t // 30) % len(patrol_route)  # 每30秒切换目标
        return patrol_route[cycle_pos]

    best = None
    best_score = -1
    for idx, f_r, f_c, f_v, rem in active:
        d = abs(robot_r - f_r) + abs(robot_c - f_c)
        if d <= rem:
            score = f_v / (d + 0.5)
            if score > best_score:
                best_score = score
                best = (f_r, f_c)
    if best:
        return best
    # 无可达食物，继续巡逻
    cycle_pos = (t // 30) % len(patrol_route)
    return patrol_route[cycle_pos]


# 17c. 改进的自适应策略：考虑到达食物后的位置价值
# 使用预计算的价值图 + 即时奖励
def strategy_adaptive_v2(robot_r, robot_c, active, t):
    """策略I: 自适应V2 - 改进的等待位置选择与食物优先级"""
    if not active:
        # 只在局部范围（距离≤5）内寻找最佳等待位置
        best_pos = (robot_r, robot_c)
        best_val = expected_value_map[robot_r, robot_c]
        for r in range(max(0, robot_r-5), min(GRID_SIZE, robot_r+6)):
            for c in range(max(0, robot_c-5), min(GRID_SIZE, robot_c+6)):
                d = abs(robot_r - r) + abs(robot_c - c)
                remaining_fraction = (TOTAL_TIME - t - d) / TOTAL_TIME
                if remaining_fraction <= 0:
                    continue
                # 期望收益 = 该位置每秒期望积分 * 剩余时间 - 移动时间的机会成本
                ev_at_pos = expected_value_map[r, c] * remaining_fraction * TOTAL_TIME
                # 移动期间错过食物的机会成本
                opportunity_cost = d * expected_value_map[robot_r, robot_c] * 0.3
                val = ev_at_pos - opportunity_cost
                if val > best_val:
                    best_val = val
                    best_pos = (r, c)
        return best_pos

    # 有食物时：选可达的价值最高的
    best = None
    best_score = -1
    for idx, f_r, f_c, f_v, rem in active:
        d = abs(robot_r - f_r) + abs(robot_c - f_c)
        if d <= rem:
            score = f_v / (d + 0.5)  # 价值/距离权衡
            if score > best_score:
                best_score = score
                best = (f_r, f_c)

    if best:
        return best
    # 无可达食物，保持原位
    return (robot_r, robot_c)




def run_full_analysis():
    """Full EDA + strategy comparison + bootstrap analysis (the original
    script body). Run `python3 learn.py` to execute it; importing learn
    only loads the data, maps and strategy functions above."""
    # ============================================================
    # 6. 多食物并发分析
    # ============================================================
    print("\n" + "=" * 60)
    print("6. 多食物并发分析")
    print("=" * 60)

    # 计算每个时刻活跃的食物数量
    active_counts = []
    for t in range(0, TOTAL_TIME, 1):
        active = np.sum((times >= t) & (times < t + FOOD_LIFETIME))
        active_counts.append(active)

    active_counts = np.array(active_counts)
    print(f"平均同时活跃食物数: {active_counts.mean():.2f}")
    print(f"最大同时活跃食物数: {active_counts.max():.0f}")
    print(f"无食物时段占比: {(active_counts==0).mean()*100:.1f}%")
    print(f"仅1个食物时段占比: {(active_counts==1).mean()*100:.1f}%")
    print(f"≥2个食物时段占比: {(active_counts>=2).mean()*100:.1f}%")
    print(f"≥3个食物时段占比: {(active_counts>=3).mean()*100:.1f}%")

    # ============================================================
    # 7. 可视化
    # ============================================================
    print("\n" + "=" * 60)
    print("7. 生成可视化图表")
    print("=" * 60)

    fig, axes = plt.subplots(2, 3, figsize=(18, 12))

    # 7a. 食物出现时间序列 (前500秒)
    ax = axes[0, 0]
    sample_t = times[times <= 500]
    sample_v = values[:len(sample_t)]
    ax.scatter(sample_t, sample_v, alpha=0.6, s=20, c='steelblue')
    ax.set_xlabel('Time (s)')
    ax.set_ylabel('Value')
    ax.set_title('Food Appearance Time Series (first 500s)')
    ax.axhline(y=values.mean(), color='red', linestyle='--', label=f'Mean={values.mean():.1f}')
    ax.legend()

    # 7b. 分值分布直方图
    ax = axes[0, 1]
    ax.hist(values, bins=40, color='steelblue', edgecolor='white', alpha=0.8)
    ax.axvline(x=values.mean(), color='red', linestyle='--', label=f'Mean={values.mean():.1f}')
    ax.set_xlabel('Value')
    ax.set_ylabel('Frequency')
    ax.set_title('Food Value Distribution')
    ax.legend()

    # 7c. 到达间隔分布
    ax = axes[0, 2]
    ax.hist(inter_arrival, bins=13, color='coral', edgecolor='white', alpha=0.8)
    ax.set_xlabel('Interval (s)')
    ax.set_ylabel('Frequency')
    ax.set_title('Food Inter-arrival Time Distribution')
    ax.axvline(x=inter_arrival.mean(), color='red', linestyle='--', label=f'Mean={inter_arrival.mean():.1f}s')
    ax.legend()

    # 7d. 空间热力图 - 出现次数
    ax = axes[1, 0]
    im1 = ax.imshow(grid_count, cmap='YlOrRd', origin='lower')
    ax.set_title('Food Count Heatmap')
    ax.set_xlabel('Column')
    ax.set_ylabel('Row')
    ax.set_xticks(range(GRID_SIZE))
    ax.set_yticks(range(GRID_SIZE))
    ax.set_xticklabels(range(1, GRID_SIZE+1))
    ax.set_yticklabels(range(1, GRID_SIZE+1))
    for r in range(GRID_SIZE):
        for c in range(GRID_SIZE):
            ax.text(c, r, int(grid_count[r, c]), ha='center', va='center', fontsize=8)
    plt.colorbar(im1, ax=ax)

    # 7e. 空间热力图 - 平均分值
    ax = axes[1, 1]
    im2 = ax.imshow(grid_avg_value, cmap='YlGnBu', origin='lower')
    ax.set_title('Average Value Heatmap')
    ax.set_xlabel('Column')
    ax.set_ylabel('Row')
    ax.set_xticks(range(GRID_SIZE))
    ax.set_yticks(range(GRID_SIZE))
    ax.set_xticklabels(range(1, GRID_SIZE+1))
    ax.set_yticklabels(range(1, GRID_SIZE+1))
    for r in range(GRID_SIZE):
        for c in range(GRID_SIZE):
            if grid_count[r, c] > 0:
                ax.text(c, r, f'{grid_avg_value[r,c]:.1f}', ha='center', va='center', fontsize=7)
    plt.colorbar(im2, ax=ax)

    # 7f. 覆盖总积分热力图 (3步可达)
    ax = axes[1, 2]
    im3 = ax.imshow(coverage_value, cmap='RdYlGn', origin='lower')
    ax.set_title('Total Reachable Value (≤3 steps)')
    ax.set_xlabel('Column')
    ax.set_ylabel('Row')
    ax.set_xticks(range(GRID_SIZE))
    ax.set_yticks(range(GRID_SIZE))
    ax.set_xticklabels(range(1, GRID_SIZE+1))
    ax.set_yticklabels(range(1, GRID_SIZE+1))
    flat_idx = np.argsort(coverage_value.flatten())[::-1][:3]
    for i, idx in enumerate(flat_idx):
        r, c = idx // GRID_SIZE, idx % GRID_SIZE
        ax.scatter(c, r, s=200, c='blue', marker='*')
        ax.annotate(f'#{i+1}', (c+0.3, r+0.3), fontsize=10, color='blue', fontweight='bold')
    plt.colorbar(im3, ax=ax)

    plt.tight_layout()
    plt.savefig('eda_analysis.png', dpi=150, bbox_inches='tight')
    print("已保存 eda_analysis.png")

    print("\nEDA 完成!")

    # ============================================================
    # 8. 食物出现规律建模 (ML)
    # ============================================================
    print("\n" + "=" * 60)
    print("8. 机器学习建模 - 食物出现规律")
    print("=" * 60)

    # 8a. 时间维度：到达过程建模
    # 检验是否服从泊松过程
    from scipy import stats

    # 每分钟食物数量
    minute_counts = []
    for m in range(180):  # 180 minutes in 3 hours
        mask = (times >= m*60) & (times < (m+1)*60)
        minute_counts.append(mask.sum())

    minute_counts = np.array(minute_counts)
    print(f"\n每分钟食物数: mean={minute_counts.mean():.2f}, var={minute_counts.var():.2f}")
    print(f"泊松分布检验: 均值≈方差 → {'符合' if abs(minute_counts.mean()-minute_counts.var()) < 2 else '略有过离散'}")

    # 8b. 空间维度：多项式分布拟合
    # 计算每个位置的出现概率
    grid_prob = grid_count / grid_count.sum()
    print(f"\n最高概率位置: ({grid_prob.argmax()//10+1}, {grid_prob.argmax()%10+1}), p={grid_prob.max():.4f}")

    # 8c. 分值分布拟合
    # 尝试拟合指数分布或幂律分布
    from scipy.optimize import curve_fit

    # 分值频率
    val_freq = Counter(values)
    vals_sorted = sorted(val_freq.keys())
    freqs = [val_freq[v] for v in vals_sorted]

    # 幂律拟合: P(v) ∝ v^(-alpha)
    def power_law(x, a, c):
        return c * x**(-a)

    try:
        popt, _ = curve_fit(power_law, np.array(vals_sorted[1:]), np.array(freqs[1:]), maxfev=5000)
        print(f"\n分值幂律分布拟合: P(v) ∝ v^(-{popt[0]:.2f})")
    except:
        print("\n分值分布: 近似指数衰减")

    # 8d. 条件概率模型：给定当前位置，下一食物的位置分布
    # 分析食物出现的空间自相关性
    print("\n食物序列空间相关性分析:")
    # 计算连续食物之间的空间距离分布
    consec_dists = []
    for i in range(N-1):
        d = manhattan(rows_arr[i]-1, cols_arr[i]-1, rows_arr[i+1]-1, cols_arr[i+1]-1)
        consec_dists.append(d)
    consec_dists = np.array(consec_dists)
    print(f"  连续食物间距离: mean={consec_dists.mean():.2f}, 独立分布期望=5.5")
    print(f"  → 食物位置{'有' if abs(consec_dists.mean()-5.5) > 0.3 else '无明显'}空间自相关")

    # 8e. 时间-空间联合分布
    # 将3小时分成6个时段，看空间分布是否变化
    print("\n时段空间分布变化 (卡方检验):")
    time_segments = [(0, 1800), (1800, 3600), (3600, 5400), (5400, 7200), (7200, 9000), (9000, 10800)]
    for i, (start, end) in enumerate(time_segments):
        mask = (times >= start) & (times < end)
        seg_grid = np.zeros((GRID_SIZE, GRID_SIZE))
        for j in np.where(mask)[0]:
            seg_grid[rows_arr[j]-1, cols_arr[j]-1] += 1
        # 与全局分布的相关系数
        corr = np.corrcoef(grid_count.flatten(), seg_grid.flatten())[0, 1]
        print(f"  时段{i+1} ({start//60}-{end//60}min): 与全局相关系数={corr:.4f}")

    # ============================================================
    # 11. 运行仿真对比
    # ============================================================
    print("\n" + "=" * 60)
    print("11. 仿真对比 - 所有策略")
    print("=" * 60)

    sim = RobotSimulator(times, rows_arr-1, cols_arr-1, values,
                         food_lifetime=FOOD_LIFETIME, total_time=TOTAL_TIME)

    strategies = {
        'A-静态最佳位置(3,7)': strategy_static_best,
        'B-贪心最高价值': strategy_greedy_value,
        'C-价值距离比最优': strategy_value_per_distance,
        'D-自适应等待': strategy_adaptive_wait,
        'E-ML最优价值图': strategy_ml_optimal,
    }

    results = {}
    for name, strat_fn in strategies.items():
        print(f"\n仿真: {name}...")
        result = sim.simulate_discrete(strat_fn, start_pos=(0, 0))  # (0,0) = (1,1)
        results[name] = result
        print(f"  总分: {result['total_score']}")
        print(f"  吃到的食物: {result['eaten_count']} / {N}")
        print(f"  每分钟积分: {result['score_per_minute']:.2f}")
        print(f"  每小时积分: {result['score_per_hour']:.0f}")

    # ============================================================
    # 12. 进一步优化 - 蒙特卡洛参数搜索
    # ============================================================
    print("\n" + "=" * 60)
    print("12. 蒙特卡洛参数优化")
    print("=" * 60)

    # 网格搜索最优参数
    print("搜索最优 (alpha, beta) 参数...")
    best_params = None
    best_param_score = 0
    param_scores = []

    for alpha in [0.5, 0.8, 1.0, 1.2, 1.5]:
        for beta in [0.5, 1.0, 1.5, 2.0]:
            strat = make_weighted_strategy(alpha, beta)
            result = sim.simulate_discrete(strat, start_pos=(0, 0))
            param_scores.append((alpha, beta, result['total_score']))
            if result['total_score'] > best_param_score:
                best_param_score = result['total_score']
                best_params = (alpha, beta)
            print(f"  alpha={alpha:.1f}, beta={beta:.1f}: {result['total_score']}分")

    print(f"\n最优参数: alpha={best_params[0]}, beta={best_params[1]}, 得分={best_param_score}")

    # 最优参数策略仿真
    strategy_optimized = make_weighted_strategy(best_params[0], best_params[1])
    opt_result = sim.simulate_discrete(strategy_optimized, start_pos=(0, 0))
    results['F-参数优化策略'] = opt_result
    print(f"  优化策略: 总分={opt_result['total_score']}, 每分钟={opt_result['score_per_minute']:.2f}")

    # ============================================================
    # 13. 结果汇总与可视化
    # ============================================================
    print("\n" + "=" * 60)
    print("13. 最终结果汇总")
    print("=" * 60)

    print(f"\n{'策略':<25s} {'总分':>8s} {'食物数':>6s} {'分/分钟':>8s} {'分/小时':>8s}")
    print("-" * 60)
    for name, r in results.items():
        print(f"{name:<25s} {r['total_score']:>8d} {r['eaten_count']:>6d} {r['score_per_minute']:>8.2f} {r['score_per_hour']:>8.0f}")

    # 理论最大值：如果所有食物都能被吃到
    theoretical_max = values.sum()
    print(f"\n理论最大总分(吃到所有食物): {theoretical_max}")
    print(f"最优策略捕获率: {max(r['total_score'] for r in results.values())/theoretical_max*100:.1f}%")

    # ============================================================
    # 14. 可视化结果
    # ============================================================
    print("\n" + "=" * 60)
    print("14. 结果可视化")
    print("=" * 60)

    fig2, axes2 = plt.subplots(2, 2, figsize=(14, 12))

    # 14a. 策略对比柱状图
    ax = axes2[0, 0]
    names_short = [n.split('-')[0] for n in results.keys()]
    scores = [r['total_score'] for r in results.values()]
    colors = plt.cm.viridis(np.linspace(0.2, 0.9, len(scores)))
    bars = ax.bar(range(len(scores)), scores, color=colors)
    ax.set_xticks(range(len(scores)))
    ax.set_xticklabels(names_short)
    ax.set_ylabel('Total Score')
    ax.set_title('Strategy Comparison - Total Score (3 hours)')
    for bar, score in zip(bars, scores):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 100, str(score),
                ha='center', va='bottom', fontweight='bold', fontsize=9)

    # 14b. 每分钟积分对比
    ax = axes2[0, 1]
    spm = [r['score_per_minute'] for r in results.values()]
    bars = ax.bar(range(len(spm)), spm, color=colors)
    ax.set_xticks(range(len(spm)))
    ax.set_xticklabels(names_short)
    ax.set_ylabel('Score / Minute')
    ax.set_title('Strategy Comparison - Score per Minute')
    for bar, s in zip(bars, spm):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.05, f'{s:.2f}',
                ha='center', va='bottom', fontweight='bold', fontsize=9)

    # 14c. 最优参数热力图
    ax = axes2[1, 0]
    alphas = sorted(set(p[0] for p in param_scores))
    betas = sorted(set(p[1] for p in param_scores))
    heatmap = np.zeros((len(alphas), len(betas)))
    for a, b, s in param_scores:
        ai = alphas.index(a)
        bi = betas.index(b)
        heatmap[ai, bi] = s
    im = ax.imshow(heatmap, cmap='YlOrRd', aspect='auto', origin='lower')
    ax.set_xticks(range(len(betas)))
    ax.set_xticklabels([f'{b:.1f}' for b in betas])
    ax.set_yticks(range(len(alphas)))
    ax.set_yticklabels([f'{a:.1f}' for a in alphas])
    ax.set_xlabel('beta (distance penalty)')
    ax.set_ylabel('alpha (value weight)')
    ax.set_title('Parameter Optimization Heatmap')
    for ai in range(len(alphas)):
        for bi in range(len(betas)):
            ax.text(bi, ai, int(heatmap[ai, bi]), ha='center', va='center', fontsize=8)
    plt.colorbar(im, ax=ax)

    # 14d. 最优等待位置可视化
    ax = axes2[1, 1]
    im = ax.imshow(optimal_value_map, cmap='RdYlGn', origin='lower')
    ax.set_title('Position Value Map (for waiting strategy)')
    ax.set_xlabel('Column')
    ax.set_ylabel('Row')
    ax.set_xticks(range(GRID_SIZE))
    ax.set_yticks(range(GRID_SIZE))
    ax.set_xticklabels(range(1, GRID_SIZE+1))
    ax.set_yticklabels(range(1, GRID_SIZE+1))
    best_r, best_c = np.unravel_index(optimal_value_map.argmax(), optimal_value_map.shape)
    ax.scatter(best_c, best_r, s=300, c='blue', marker='*')
    ax.annotate(f'Best\n({best_r+1},{best_c+1})', (best_c, best_r+0.5),
                fontsize=10, color='blue', fontweight='bold', ha='center')
    plt.colorbar(im, ax=ax)

    plt.tight_layout()
    plt.savefig('strategy_results.png', dpi=150, bbox_inches='tight')
    print("已保存 strategy_results.png")

    # ============================================================
    # 15. 敏感性分析与可行性讨论
    # ============================================================
    print("\n" + "=" * 60)
    print("15. 敏感性分析")
    print("=" * 60)

    # 15a. 食物生命周期变化的影响
    print("\n食物生命周期敏感性:")
    for lt in [2, 3, 4, 5]:
        sim_alt = RobotSimulator(times, rows_arr-1, cols_arr-1, values,
                                 food_lifetime=lt, total_time=TOTAL_TIME)
        r = sim_alt.simulate_discrete(strategy_optimized, start_pos=(0, 0))
        print(f"  生命周期={lt}s: 总分={r['total_score']}, 分/分钟={r['score_per_minute']:.2f}")

    # 15b. 起始位置影响
    print("\n起始位置敏感性:")
    for start_pos in [(0,0), (4,4), (2,6), (7,2), (4,7)]:
        r = sim.simulate_discrete(strategy_optimized, start_pos=start_pos)
        sr, sc = start_pos
        print(f"  起始({sr+1},{sc+1}): 总分={r['total_score']}, 分/分钟={r['score_per_minute']:.2f}")

    # 15c. 实际可行性分析
    print("\n" + "=" * 60)
    print("16. 实际竞赛可行性分析")
    print("=" * 60)

    best_result = max(results.values(), key=lambda x: x['total_score'])
    print(f"""
    可行性分析:
    -----------
    1. 策略效果: 3小时内预期获得 {best_result['total_score']} 分
    2. 每分钟积分: {best_result['score_per_minute']:.2f} 分/分钟
    3. 每小时积分: {best_result['score_per_hour']:.0f} 分/小时
    4. 食物捕获率: {best_result['eaten_count']/N*100:.1f}% ({best_result['eaten_count']}/{N})
    5. 理论最大捕获率(距离约束): 因为食物存活仅3秒，机器人从其覆盖范围
       内可捕获的食物比例上限取决于等待位置。最佳位置(3,7)覆盖{int(coverage_count[best_wait_pos])}个食物
       (占全部的{coverage_count[best_wait_pos]/N*100:.1f}%)

    实际应用建议:
    - 机器人应优先部署在场地中部的热区(第7-9列)
    - 当多个食物同时出现时，优先选择价值高的
    - 空闲时应在(3,7)附近巡逻等待
    - 实际比赛中食物模式可能与历史数据有差异，建议保留安全边际
    - 如有实时学习能力，可在比赛中动态调整策略

    局限性:
    - 历史数据可能不完全代表实际比赛的食物生成规律
    - 3秒约束极为严格，机器人响应速度和加速度需达标
    - 实际比赛中可能存在通信延迟、定位误差等因素
    """)


    r_time = sim.simulate_discrete(strategy_time_aware, start_pos=(0, 0))
    results['G-时间感知策略'] = r_time
    print(f"  时间感知策略: 总分={r_time['total_score']}, 分/分钟={r_time['score_per_minute']:.2f}")

    r_patrol = sim.simulate_discrete(strategy_patrol, start_pos=(0, 0))
    results['H-巡逻策略'] = r_patrol
    print(f"  巡逻策略: 总分={r_patrol['total_score']}, 分/分钟={r_patrol['score_per_minute']:.2f}")

    r_adaptive2 = sim.simulate_discrete(strategy_adaptive_v2, start_pos=(0, 0))
    results['I-自适应V2'] = r_adaptive2
    print(f"  自适应V2: 总分={r_adaptive2['total_score']}, 分/分钟={r_adaptive2['score_per_minute']:.2f}")

    # ============================================================
    # 18. 最优策略深度分析 - 交易曲面
    # ============================================================
    print("\n" + "=" * 60)
    print("18. 深度分析 - 最优策略特征")
    print("=" * 60)

    # 18a. 什么特征决定了食物是否被吃到？
    print("\n分析食物被捕获的特征...")
    # 用最优策略D模拟结果
    best_strat_for_analysis = strategy_adaptive_wait
    detailed_result = sim.simulate_discrete(best_strat_for_analysis, start_pos=(0, 0))

    # 追踪哪些食物被吃了
    eaten_set = set()
    robot_r, robot_c = 0, 0
    target_r, target_c = 0, 0
    food_eaten_at = {}  # food_idx -> time eaten

    # 重新模拟并记录
    for t in range(TOTAL_TIME + 1):
        active = []
        for i, (f_t, f_r, f_c, f_v) in enumerate(sim.foods):
            if i in eaten_set:
                continue
            if f_t <= t < f_t + FOOD_LIFETIME:
                remaining = FOOD_LIFETIME - (t - f_t)
                active.append((i, f_r, f_c, f_v, remaining))

        # 检查是否在食物上
        for i, f_r, f_c, f_v, rem in active:
            if robot_r == f_r and robot_c == f_c and i not in eaten_set:
                eaten_set.add(i)
                food_eaten_at[i] = t

        # 决策
        target = best_strat_for_analysis(robot_r, robot_c, active, t)
        if target is not None:
            target_r, target_c = target

        # 移动
        if robot_r < target_r:
            robot_r += 1
        elif robot_r > target_r:
            robot_r -= 1
        elif robot_c < target_c:
            robot_c += 1
        elif robot_c > target_c:
            robot_c -= 1

    # 分析被吃食物 vs 未被吃食物的特征
    eaten_vals = [sim.foods[i][3] for i in eaten_set]
    missed_vals = [sim.foods[i][3] for i in range(N) if i not in eaten_set]
    print(f"被吃食物平均分值: {np.mean(eaten_vals):.2f}, 中位数: {np.median(eaten_vals):.0f}")
    print(f"错过食物平均分值: {np.mean(missed_vals):.2f}, 中位数: {np.median(missed_vals):.0f}")

    # 空间分析
    eaten_grid = np.zeros((GRID_SIZE, GRID_SIZE))
    missed_grid = np.zeros((GRID_SIZE, GRID_SIZE))
    for i in eaten_set:
        f_t, f_r, f_c, f_v = sim.foods[i]
        eaten_grid[f_r, f_c] += 1
    for i in range(N):
        if i not in eaten_set:
            f_t, f_r, f_c, f_v = sim.foods[i]
            missed_grid[f_r, f_c] += 1

    print("\n被吃食物热力图:")
    for r in range(GRID_SIZE):
        print(' '.join(f'{int(eaten_grid[r,c]):3d}' for c in range(GRID_SIZE)))

    print("\n捕获率热力图 (%):")
    for r in range(GRID_SIZE):
        row_str = ' '.join(f'{eaten_grid[r,c]/grid_count[r,c]*100:5.0f}' if grid_count[r,c]>0 else '   0' for c in range(GRID_SIZE))
        print(f"  行{r+1}: {row_str}")

    # 18b. 时间分布分析
    print("\n按时段捕获率:")
    for h in range(3):
        start, end = h*3600, (h+1)*3600
        eaten_in_hour = sum(1 for i in eaten_set if start <= sim.foods[i][0] < end)
        total_in_hour = sum(1 for i in range(N) if start <= sim.foods[i][0] < end)
        print(f"  第{h+1}小时: {eaten_in_hour}/{total_in_hour} = {eaten_in_hour/total_in_hour*100:.1f}%")

    # ============================================================
    # 19. 机器学习模型：预测食物可捕获性
    # ============================================================
    print("\n" + "=" * 60)
    print("19. 机器学习 - 预测食物可捕获性")
    print("=" * 60)

    # 构建特征矩阵
    features = []
    labels = []
    for i in range(N):
        f_t, f_r, f_c, f_v = sim.foods[i]
        # 特征
        feat = [
            f_t, f_r, f_c, f_v,  # 基本特征
            f_t % 60,  # 秒
            (f_t // 60) % 60,  # 分
            f_t // 3600,  # 时
            f_r + f_c,  # 曼哈顿距离从(1,1)
            abs(f_r - 5.5) + abs(f_c - 5.5),  # 到中心的距离
            grid_count[f_r, f_c],  # 该位置历史频率
            grid_avg_value[f_r, f_c],  # 该位置平均分值
        ]
        # 检查前一个食物
        if i > 0:
            prev_t, prev_r, prev_c, prev_v = sim.foods[i-1]
            feat.extend([
                f_t - prev_t,  # 间隔
                abs(f_r - prev_r) + abs(f_c - prev_c),  # 空间距离
                prev_v,  # 前一个分值
            ])
        else:
            feat.extend([0, 0, 0])

        features.append(feat)
        labels.append(1 if i in eaten_set else 0)

    X = np.array(features)
    y = np.array(labels)

    print(f"特征矩阵: {X.shape}")
    print(f"正样本(被吃): {y.sum()}, 负样本(错过): {(1-y).sum()}")

    # 训练XGBoost
    try:
        import xgboost as xgb
        from sklearn.model_selection import train_test_split
        from sklearn.metrics import classification_report, roc_auc_score

        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
        model = xgb.XGBClassifier(n_estimators=100, max_depth=5, learning_rate=0.1, random_state=42)
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        y_prob = model.predict_proba(X_test)[:, 1]
        print(f"\nXGBoost分类报告:")
        print(f"  准确率: {(y_pred == y_test).mean():.4f}")
        print(f"  AUC: {roc_auc_score(y_test, y_prob):.4f}")

        # 特征重要性
        importance = model.feature_importances_
        feat_names = ['time', 'row', 'col', 'value', 'sec', 'min', 'hour',
                      'dist_from_start', 'dist_from_center', 'pos_freq', 'pos_avg_val',
                      'interval', 'spatial_dist', 'prev_value']
        print("\n特征重要性 Top 5:")
        for idx in np.argsort(importance)[::-1][:5]:
            print(f"  {feat_names[idx]}: {importance[idx]:.4f}")
    except ImportError:
        print("  XGBoost未安装，跳过ML模型训练")

    # ============================================================
    # 20. 引导式仿真：Bootstrap评估策略稳定性
    # ============================================================
    print("\n" + "=" * 60)
    print("20. Bootstrap稳定性评估")
    print("=" * 60)

    # 使用Block Bootstrap保持时间结构
    n_bootstrap = 50
    bootstrap_scores = []
    block_duration = 1800  # 30分钟块
    n_blocks = TOTAL_TIME // block_duration  # 6 blocks

    for b in range(n_bootstrap):
        # 随机抽取块（有放回）
        selected_blocks = np.random.choice(n_blocks, size=n_blocks, replace=True)
        bs_foods = []
        for blk in selected_blocks:
            start_t = blk * block_duration
            end_t = (blk + 1) * block_duration
            mask = (times >= start_t) & (times < end_t)
            offset = len(bs_foods) * 0  # 保持原始时间
            for i in np.where(mask)[0]:
                bs_foods.append((times[i], rows_arr[i]-1, cols_arr[i]-1, values[i]))

        bs_foods.sort(key=lambda x: x[0])
        if len(bs_foods) == 0:
            continue
        bs_times = np.array([f[0] for f in bs_foods])
        bs_rows = np.array([f[1] for f in bs_foods])
        bs_cols = np.array([f[2] for f in bs_foods])
        bs_values = np.array([f[3] for f in bs_foods])

        bs_sim = RobotSimulator(bs_times, bs_rows, bs_cols, bs_values,
                                food_lifetime=FOOD_LIFETIME, total_time=TOTAL_TIME)
        # 用最佳策略G
        bs_result = bs_sim.simulate_discrete(strategy_time_aware, start_pos=(0, 0))
        bootstrap_scores.append(bs_result['total_score'])

    bootstrap_scores = np.array(bootstrap_scores)
    print(f"Bootstrap ({n_bootstrap}次):")
    print(f"  平均总分: {bootstrap_scores.mean():.0f} ± {bootstrap_scores.std():.0f}")
    print(f"  95% CI: [{np.percentile(bootstrap_scores, 2.5):.0f}, {np.percentile(bootstrap_scores, 97.5):.0f}]")
    print(f"  平均分/分钟: {bootstrap_scores.mean()/180:.2f} ± {bootstrap_scores.std()/180:.2f}")

    # ============================================================
    # 21. 最终结果汇总
    # ============================================================
    print("\n" + "=" * 60)
    print("21. 最终结果汇总")
    print("=" * 60)

    print(f"\n{'策略':<30s} {'总分':>8s} {'食物数':>6s} {'分/分钟':>8s} {'捕获率':>6s}")
    print("-" * 70)
    for name, r in sorted(results.items(), key=lambda x: x[1]['total_score'], reverse=True):
        cap_rate = r['eaten_count'] / N * 100
        print(f"{name:<30s} {r['total_score']:>8d} {r['eaten_count']:>6d} {r['score_per_minute']:>8.2f} {cap_rate:>5.1f}%")

    best_result = max(results.values(), key=lambda x: x['total_score'])
    best_name = max(results.items(), key=lambda x: x[1]['total_score'])[0]
    print(f"\n*** 最优策略: {best_name}")
    print(f"   总分: {best_result['total_score']}")
    print(f"   每分钟: {best_result['score_per_minute']:.2f}")
    print(f"   每小时: {best_result['score_per_hour']:.0f}")
    print(f"   捕获率: {best_result['eaten_count']/N*100:.1f}%")

    # ============================================================
    # 22. 最终可视化
    # ============================================================
    fig3, axes3 = plt.subplots(2, 3, figsize=(18, 12))

    # 22a. 所有策略总分对比
    ax = axes3[0, 0]
    names = [n.replace('策略', '\n策略') for n in results.keys()]
    scores = [r['total_score'] for r in results.values()]
    colors = plt.cm.RdYlGn(np.linspace(0.2, 0.9, len(scores)))
    bars = ax.barh(range(len(scores)), scores, color=colors)
    ax.set_yticks(range(len(scores)))
    ax.set_yticklabels([n[:25] for n in results.keys()], fontsize=8)
    ax.set_xlabel('Total Score (3 hours)')
    ax.set_title('All Strategies - Total Score Comparison')
    for i, (bar, score) in enumerate(zip(bars, scores)):
        ax.text(bar.get_width() + 30, bar.get_y() + bar.get_height()/2,
                str(score), va='center', fontweight='bold')

    # 22b. Bootstrap分布
    ax = axes3[0, 1]
    ax.hist(bootstrap_scores, bins=20, color='steelblue', edgecolor='white', alpha=0.8)
    ax.axvline(x=bootstrap_scores.mean(), color='red', linestyle='-', linewidth=2, label=f'Mean={bootstrap_scores.mean():.0f}')
    ax.axvline(x=np.percentile(bootstrap_scores, 2.5), color='orange', linestyle='--', label='95% CI')
    ax.axvline(x=np.percentile(bootstrap_scores, 97.5), color='orange', linestyle='--')
    ax.set_xlabel('Total Score')
    ax.set_ylabel('Frequency')
    ax.set_title('Bootstrap Score Distribution')
    ax.legend()

    # 22c. 捕获率热力图
    ax = axes3[0, 2]
    capture_rate_grid = np.divide(eaten_grid, grid_count, where=grid_count>0) * 100
    im = ax.imshow(capture_rate_grid, cmap='RdYlGn', origin='lower', vmin=0, vmax=100)
    ax.set_title('Capture Rate by Position (%)')
    for r in range(GRID_SIZE):
        for c in range(GRID_SIZE):
            if grid_count[r, c] > 0:
                ax.text(c, r, f'{capture_rate_grid[r,c]:.0f}', ha='center', va='center', fontsize=7)
    plt.colorbar(im, ax=ax)

    # 22d. 食物出现 vs 捕获散点图
    ax = axes3[1, 0]
    ax.scatter(rows_arr, cols_arr, c='lightgray', s=10, alpha=0.3, label='All Foods')
    eaten_rows = [sim.foods[i][1] for i in eaten_set]
    eaten_cols = [sim.foods[i][2] for i in eaten_set]
    ax.scatter(eaten_rows, eaten_cols, c='green', s=15, alpha=0.7, label=f'Eaten (n={len(eaten_set)})')
    ax.scatter([0], [0], c='red', s=100, marker='*', label='Start (1,1)')
    wr, wc = best_wait_pos
    ax.scatter([wr], [wc], c='blue', s=100, marker='s', label=f'Best Wait ({wr+1},{wc+1})')
    ax.set_xlabel('Row')
    ax.set_ylabel('Column')
    ax.set_title('Food Positions: Eaten vs Missed')
    ax.set_xlim(-1, 11)
    ax.set_ylim(-1, 11)
    ax.legend(fontsize=8)

    # 22e. 生命周期敏感性
    ax = axes3[1, 1]
    lifetimes = [2, 3, 4, 5, 6, 7]
    lt_scores = []
    for lt in lifetimes:
        sim_alt = RobotSimulator(times, rows_arr-1, cols_arr-1, values,
                                 food_lifetime=lt, total_time=TOTAL_TIME)
        r = sim_alt.simulate_discrete(strategy_adaptive_v2, start_pos=(0, 0))
        lt_scores.append(r['total_score'])
    ax.plot(lifetimes, lt_scores, 'o-', color='steelblue', linewidth=2, markersize=10)
    ax.set_xlabel('Food Lifetime (seconds)')
    ax.set_ylabel('Total Score')
    ax.set_title('Score vs Food Lifetime')
    for lt, s in zip(lifetimes, lt_scores):
        ax.annotate(str(s), (lt, s), textcoords="offset points", xytext=(0, 10), ha='center')

    # 22f. 最优策略得分拆解
    ax = axes3[1, 2]
    # 按分值区间拆解
    val_bins_plot = [0, 5, 10, 15, 20, 25, 30, 41]
    eaten_by_bin = []
    total_by_bin = []
    for i in range(len(val_bins_plot)-1):
        mask = (values >= val_bins_plot[i]) & (values < val_bins_plot[i+1])
        total_by_bin.append(mask.sum())
        eaten_mask = np.array([idx in eaten_set for idx in np.where(mask)[0]])
        eaten_in_range = sum(1 for j in np.where(mask)[0] if j in eaten_set)
        eaten_by_bin.append(eaten_in_range)

    x = np.arange(len(val_bins_plot)-1)
    width = 0.35
    ax.bar(x - width/2, total_by_bin, width, label='Total Foods', color='lightgray')
    ax.bar(x + width/2, eaten_by_bin, width, label='Eaten', color='green')
    ax.set_xticks(x)
    ax.set_xticklabels([f'[{val_bins_plot[i]},{val_bins_plot[i+1]})' for i in range(len(val_bins_plot)-1)], fontsize=7)
    ax.set_ylabel('Count')
    ax.set_title('Foods by Value Range: Total vs Eaten')
    ax.legend()

    plt.tight_layout()
    plt.savefig('advanced_analysis.png', dpi=150, bbox_inches='tight')
    print("\n已保存 advanced_analysis.png")

    print("\n" + "=" * 60)
    print("全部分析完成! 请查看以下文件:")
    print("  eda_analysis.png - 探索性数据分析")
    print("  strategy_results.png - 策略对比结果")
    print("  advanced_analysis.png - 高级分析")
    print("=" * 60)


if __name__ == '__main__':
    run_full_analysis()
