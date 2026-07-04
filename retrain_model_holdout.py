"""Same modeling approach as retrain_model.py (behavior-cloned XGBoost /
RandomForest / MLP imitating a time-block-aware teacher strategy), but
built ONLY from the first N_TRAIN events (chronologically) so it can be
evaluated on a genuinely held-out test set, the way strategy_blackbox.py's
--train-events already allows for the hand-designed strategy.

Every statistic here (grid coverage, per-block best waiting position, the
behavior-cloning replay used to train the classifier) is computed from
events[:N_TRAIN] alone. Time features are relative to the training
window's own start (t - T_START), so testenv/retrain_holdout_blackbox.py
must feed it t relative to when the *test* protocol starts, not the
dataset's absolute clock -- otherwise the model would see out-of-range
time features it never trained on.

Interface mirrors retrain_model.py: load_policy(), make_ml_strategies(),
strategy_time_aware(), retrain(). Model is saved to
robot_policy_model_holdout.pkl.
"""
import os
import openpyxl, numpy as np, pickle, warnings
warnings.filterwarnings('ignore')
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import xgboost as xgb
from sklearn.ensemble import RandomForestClassifier
from sklearn.neural_network import MLPClassifier

_HERE = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(_HERE, 'robot_policy_model_holdout.pkl')
N_TRAIN = 893

wb = openpyxl.load_workbook(os.path.join(_HERE, 'data.xlsx'), data_only=True)
ws = wb['Sheet1']
_data_all = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    _data_all.append(list(row))
_data_all.sort(key=lambda d: d[0])
data = _data_all[:N_TRAIN]              # training slice only -- test set untouched

times = np.array([d[0] for d in data])
values = np.array([d[2] for d in data])
rows_list, cols_list = [], []
for d in data:
    p = str(d[1]).strip('()')
    r, c = p.split(',')
    rows_list.append(int(r)-1)
    cols_list.append(int(c)-1)
rows_arr = np.array(rows_list)
cols_arr = np.array(cols_list)

GRID_SIZE, FOOD_LIFETIME = 10, 3
T_START = int(times.min())
TOTAL_TIME = int(times.max()) + FOOD_LIFETIME - T_START   # training-window length
N = len(data)
rel_times = times - T_START

grid_count = np.zeros((GRID_SIZE, GRID_SIZE))
grid_value_sum = np.zeros((GRID_SIZE, GRID_SIZE))
for i in range(N):
    r, c = rows_arr[i], cols_arr[i]
    grid_count[r, c] += 1
    grid_value_sum[r, c] += values[i]

block_size = 1800
time_blocks = max(1, -(-TOTAL_TIME // block_size))  # ceil, sized to the training window
block_best_positions = []
for blk in range(time_blocks):
    start_t, end_t = blk*block_size, (blk+1)*block_size
    mask = (rel_times >= start_t) & (rel_times < end_t)
    blk_grid_val = np.zeros((GRID_SIZE, GRID_SIZE))
    for i in np.where(mask)[0]:
        blk_grid_val[rows_arr[i], cols_arr[i]] += values[i]
    blk_coverage = np.zeros((GRID_SIZE, GRID_SIZE))
    for r in range(GRID_SIZE):
        for c in range(GRID_SIZE):
            total = 0
            for dr in range(-3, 4):
                for dc in range(-3, 4):
                    if abs(dr)+abs(dc) <= 3:
                        nr, nc = r+dr, c+dc
                        if 0 <= nr < GRID_SIZE and 0 <= nc < GRID_SIZE:
                            total += blk_grid_val[nr, nc]
            blk_coverage[r, c] = total
    if blk_coverage.max() > 0:
        block_best_positions.append(np.unravel_index(blk_coverage.argmax(), blk_coverage.shape))
    else:
        block_best_positions.append(block_best_positions[-1] if block_best_positions else (0, 0))

foods = sorted(zip(rel_times.tolist(), rows_arr.tolist(), cols_arr.tolist(), values.tolist()),
               key=lambda x: x[0])

DIRECTION_MAP = {(0,0):0, (1,0):1, (-1,0):2, (0,1):3, (0,-1):4}
REVERSE_MAP = {v:k for k,v in DIRECTION_MAP.items()}
DIRECTION_NAMES = ["(0,0)停", "(1,0)右", "(-1,0)左", "(0,1)上", "(0,-1)下"]


def build_state_vector(robot_r, robot_c, active, t):
    features = [robot_r/9.0, robot_c/9.0]
    features.append((t % block_size)/block_size)
    features.append(t/TOTAL_TIME)
    features.append((TOTAL_TIME-t)/TOTAL_TIME)
    features.append(len(active)/3.0)
    active_sorted = sorted(active, key=lambda x: x[3], reverse=True)
    for i in range(3):
        if i < len(active_sorted):
            idx, f_r, f_c, f_v, rem = active_sorted[i]
            d = abs(robot_r-f_r)+abs(robot_c-f_c)
            features.extend([f_r/9.0, f_c/9.0, f_v/40.0, rem/3.0, min(d,20)/20.0, 1.0 if d<=rem else 0.0])
        else:
            features.extend([-1.0,-1.0,0.0,0.0,1.0,0.0])
    blk = min(int(t)//block_size, time_blocks-1)
    wr, wc = block_best_positions[blk]
    features.extend([wr/9.0, wc/9.0])
    return np.array(features, dtype=np.float32)


def strategy_time_aware(robot_r, robot_c, active, t):
    blk = min(int(t)//block_size, time_blocks-1)
    wr, wc = block_best_positions[blk]
    if not active:
        return (wr, wc)
    best, best_score = None, -1
    for idx, f_r, f_c, f_v, rem in active:
        d = abs(robot_r-f_r)+abs(robot_c-f_c)
        if d <= rem:
            blk2 = min(int(t+d)//block_size, time_blocks-1)
            future_r, future_c = block_best_positions[blk2]
            future_dist = abs(f_r-future_r)+abs(f_c-future_c)
            score = f_v - 0.1*future_dist
            if score > best_score:
                best_score = score
                best = (f_r, f_c)
    return best if best else (wr, wc)


def load_policy(model_path=MODEL_PATH):
    with open(model_path, 'rb') as f:
        return pickle.load(f)


def make_ml_strategies(model_data):
    best_model = model_data['model']
    scaler = model_data['scaler']

    def ml_step(robot_r, robot_c, active, t):
        state = build_state_vector(robot_r, robot_c, active, t)
        state = scaler.transform(state.reshape(1, -1))
        action = int(best_model.predict(state)[0])
        dr, dc = REVERSE_MAP.get(action, (0, 0))
        return (max(0, min(GRID_SIZE-1, robot_r+dr)),
                max(0, min(GRID_SIZE-1, robot_c+dc)))

    def strategy_ml_hybrid(robot_r, robot_c, active, t):
        if active:
            best, best_score = None, -1
            for idx, f_r, f_c, f_v, rem in active:
                d = abs(robot_r-f_r)+abs(robot_c-f_c)
                if d <= rem:
                    score = f_v/(d+0.5)
                    if score > best_score:
                        best_score = score
                        best = (f_r, f_c)
            if best: return best
        return ml_step(robot_r, robot_c, active, t)

    def strategy_ml_pure(robot_r, robot_c, active, t):
        return ml_step(robot_r, robot_c, active, t)

    return strategy_ml_hybrid, strategy_ml_pure


def retrain(model_path=MODEL_PATH):
    print(f"Generating training data (training window only, T=0..{TOTAL_TIME})...")
    X_all, y_all = [], []
    robot_r, robot_c = 0, 0
    target_r, target_c = 0, 0
    eaten_set = set()

    for t in range(TOTAL_TIME+1):
        active = []
        for i, (f_t, f_r, f_c, f_v) in enumerate(foods):
            if i in eaten_set: continue
            if f_t <= t < f_t+FOOD_LIFETIME:
                active.append((i, f_r, f_c, f_v, FOOD_LIFETIME-(t-f_t)))
        for i, f_r, f_c, f_v, rem in active:
            if robot_r==f_r and robot_c==f_c and i not in eaten_set:
                eaten_set.add(i)
        state = build_state_vector(robot_r, robot_c, active, t)
        target = strategy_time_aware(robot_r, robot_c, active, t)
        if target is not None: target_r, target_c = target
        else: target_r, target_c = robot_r, robot_c
        dr = dc = 0
        if robot_r < target_r: dr=1; robot_r+=1
        elif robot_r > target_r: dr=-1; robot_r-=1
        elif robot_c < target_c: dc=1; robot_c+=1
        elif robot_c > target_c: dc=-1; robot_c-=1
        action = DIRECTION_MAP.get((dr,dc), 0)
        X_all.append(state)
        y_all.append(action)

    X_all = np.array(X_all, dtype=np.float32)
    y_all = np.array(y_all, dtype=np.int64)
    print(f"Total data: X={X_all.shape}, y={y_all.shape}")
    print(f"Action distribution: {dict(zip(DIRECTION_NAMES, np.bincount(y_all)))}")

    X_tr, X_te, y_tr, y_te = train_test_split(X_all, y_all, test_size=0.5, random_state=42)
    print(f"\n50/50 Split: Train={X_tr.shape[0]}, Test={X_te.shape[0]}")

    scaler = StandardScaler()
    X_tr_s = scaler.fit_transform(X_tr)
    X_te_s = scaler.transform(X_te)

    print("\n--- XGBoost ---")
    xgb_model = xgb.XGBClassifier(n_estimators=200, max_depth=8, learning_rate=0.1,
                                   subsample=0.8, colsample_bytree=0.8, random_state=42, n_jobs=-1)
    xgb_model.fit(X_tr_s, y_tr)
    xgb_acc = (xgb_model.predict(X_te_s)==y_te).mean()
    print(f"Test Accuracy (50% unseen): {xgb_acc:.4f}")

    print("\n--- RandomForest ---")
    rf_model = RandomForestClassifier(n_estimators=200, max_depth=15, random_state=42, n_jobs=-1)
    rf_model.fit(X_tr_s, y_tr)
    rf_acc = (rf_model.predict(X_te_s)==y_te).mean()
    print(f"Test Accuracy (50% unseen): {rf_acc:.4f}")

    print("\n--- MLP Neural Network ---")
    mlp_model = MLPClassifier(hidden_layer_sizes=(128,64,32), activation='relu',
                               solver='adam', max_iter=200, batch_size=256, random_state=42, verbose=False)
    mlp_model.fit(X_tr_s, y_tr)
    mlp_acc = (mlp_model.predict(X_te_s)==y_te).mean()
    print(f"Test Accuracy (50% unseen): {mlp_acc:.4f}")

    models = {"XGBoost":(xgb_model,xgb_acc), "RandomForest":(rf_model,rf_acc), "MLP":(mlp_model,mlp_acc)}
    best_name = max(models, key=lambda k: models[k][1])
    best_model, best_acc = models[best_name]
    print(f"\n*** Best Model: {best_name}, Test Accuracy={best_acc:.4f} ***")

    model_data = {
        "model": best_model, "scaler": scaler, "model_name": best_name,
        "direction_map": DIRECTION_MAP, "reverse_map": REVERSE_MAP,
        "direction_names": DIRECTION_NAMES,
        "block_best_positions": block_best_positions, "block_size": block_size,
        "time_blocks": time_blocks, "grid_size": GRID_SIZE,
        "food_lifetime": FOOD_LIFETIME, "total_time": TOTAL_TIME,
    }
    with open(model_path, "wb") as f:
        pickle.dump(model_data, f)
    print(f"Model saved to {model_path}")
    return model_data


if __name__ == '__main__':
    retrain()
